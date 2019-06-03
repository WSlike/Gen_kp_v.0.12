# Импорт сторонних библиотек
import time
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment
# Импорт внутренних файлов
from openfile import wb, f_open, save_name
import settings
import devices


"""РАБОТА С ТАБЛИЦЕЙ"""


def fill_row(bd_sheet=None, signal_type='ТС', row=0, num=None, name='-', parameter_type='-', func='-', address='-',
             ai_min='-', ai_max='-', unit='-', description='-', channel='-', comment='-'):
    """Заполнение одной строки базы данных"""

    fill = PatternFill()
    if signal_type == 'ТС':
        fill = PatternFill(fill_type='solid', start_color='9bc2e6', end_color='9bc2e6')
    elif signal_type == 'ТИ':
        fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')
    elif signal_type == 'ТУ':
        fill = PatternFill(fill_type='solid', start_color='1f4e78', end_color='1f4e78')
    elif signal_type == 'ТР':
        fill = PatternFill(fill_type='solid', start_color='c65911', end_color='c65911')

    # №
    bord_side = Side(border_style='thin', color='00000000')
    bord = Border(bottom=bord_side, left=bord_side, top=bord_side, right=bord_side)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell = bd_sheet.cell(row, 1)
    cell.fill = fill
    cell.border = bord
    cell.alignment = align
    cell.value = num

    # Наименование логического параметра
    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell = bd_sheet.cell(row, 2)
    cell.fill = fill
    cell.border = bord
    cell.alignment = align
    cell.value = name

    # Тип параметра
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell = bd_sheet.cell(row, 3)
    cell.fill = fill
    cell.border = bord
    cell.alignment = align
    cell.value = parameter_type

    # Функция ASDU
    cell = bd_sheet.cell(row, 4)
    cell.fill = fill
    cell.border = bord
    cell.alignment = align
    cell.value = func

    # Адрес объекта
    cell = bd_sheet.cell(row, 5)
    cell.fill = fill
    cell.border = bord
    cell.alignment = align
    cell.value = address

    if signal_type == 'ТС' or signal_type == 'ТУ':
        # Нижний диапазон, Верхний диапазон, Ед. измерения, Значение по умолчанию (для ТР)
        for base_k in range(4):
            cell = bd_sheet.cell(row, 6 + base_k)
            cell.fill = fill
            cell.border = bord
            cell.alignment = align
            cell.value = '-'

        # Расшифровка значения
        cell = bd_sheet.cell(row, 10)
        cell.fill = fill
        cell.border = bord
        cell.alignment = align
        cell.value = description

    elif signal_type == 'ТИ' or signal_type == 'ТР':
        # Нижний диапазон
        cell = bd_sheet.cell(row, 6)
        cell.fill = fill
        cell.border = bord
        cell.alignment = align
        cell.value = ai_min

        # Верхний диапазон
        cell = bd_sheet.cell(row, 7)
        cell.fill = fill
        cell.border = bord
        cell.alignment = align
        cell.value = ai_max

        # Ед. измерения
        cell = bd_sheet.cell(row, 8)
        cell.fill = fill
        cell.border = bord
        cell.alignment = align
        cell.value = unit

        # Значение по умолчанию (для ТР), Расшифровка значения
        for base_k in range(2):
            cell = bd_sheet.cell(row, 9 + base_k)
            cell.fill = fill
            cell.border = bord
            cell.alignment = align
            cell.value = '-'

    # № физического канала
    cell = bd_sheet.cell(row, 11)
    cell.fill = fill
    cell.border = bord
    cell.alignment = align
    cell.value = channel

    # Примечание,
    for base_k in range(2):
        cell = bd_sheet.cell(row, 12 + base_k)
        cell.fill = fill
        cell.border = bord
        cell.alignment = align
        cell.value = comment


def clr_base(bd_sheet):
    """Очищаем базу"""
    for row in range(bd_sheet.max_row, 8, -1):
        bd_sheet.delete_rows(row)


# Таймер выполнения работы
start_time = time.time()

# Если открыт файл, заполняем БД
if f_open:
    """================Заполняем БД ВУ================="""
    print('===================Заполняем БД ВУ====================')
    # задаем лист БД
    ws_bd = wb.worksheets[8]
    # очищаем базу
    clr_base(bd_sheet=ws_bd)

    """==Заголовки=="""

    print(' Заголовки')
    ws_bd['C1'] = wb.worksheets[0].cell(3, 2).value  # Название объекта телемеханизации
    modules = ''    # Расположение модулей в контроллере
    for i in range(settings.module_count):
        modules = modules + str(wb.worksheets[0].cell(10+i, 2).value) + ', '
    ws_bd['C3'] = modules
    extra1 = 'ASDU ' + str(settings.IEC_104['ASDU'])  # Дополнительная информация по контроллеру и протоколу обмена
    ws_bd['C5'] = extra1
    extra2 = 'k=' + str(settings.IEC_104['k']) + ', ' + 'w=' + str(settings.IEC_104['w']) + ', ' + 'T0=' + \
             str(settings.IEC_104['timeoutK'])
    ws_bd['L5'] = extra2

    """==Физические ТС=="""

    index = 9
    print(' Физические ТС')
    ts = wb.worksheets[1]

    for i_ts in range(settings.discrete_count):
        fill_row(bd_sheet=ws_bd,
                 signal_type='ТС',
                 row=index+i_ts,
                 num=ts.cell(6 + i_ts, 2).value,
                 name=ts.cell(6 + i_ts, 3).value,
                 parameter_type='физический',
                 func='M_SP_NA_1 (1)',
                 address=settings.IEC_104['startAddressTS'] + i_ts,
                 description=ts.cell(6 + i_ts, 13).value,
                 channel=str(ts.cell(6 + i_ts, 4).value) + "." + str(ts.cell(6 + i_ts, 5).value),
                 comment='-')

    """==Физические ТИ=="""

    print(' Физические ТИ')
    index = index + settings.discrete_count
    ti = wb.worksheets[3]

    for i_ti in range(settings.input_count):
        fill_row(bd_sheet=ws_bd,
                 signal_type='ТИ',
                 row=index+i_ti,
                 num=ti.cell(6 + i_ti, 2).value,
                 name=ti.cell(6 + i_ti, 3).value,
                 parameter_type='физический',
                 func='M_ME_NC_1 (13)',
                 address=settings.IEC_104['startAddressTI'] + i_ti,
                 ai_min=ti.cell(6 + i_ti, 10).value,
                 ai_max=ti.cell(6 + i_ti, 11).value,
                 unit=ti.cell(6 + i_ti, 17).value,
                 channel=str(ti.cell(6 + i_ti, 4).value) + "." + str(ti.cell(6 + i_ti, 5).value),
                 comment='-')

    """==Физические ТУ=="""

    print(' Физические ТУ')
    tu = wb.worksheets[2]
    index = index + settings.input_count

    for i_tu in range(settings.discrete_output_count):
        fill_row(bd_sheet=ws_bd,
                 signal_type='ТУ',
                 row=index+i_tu,
                 num=tu.cell(6 + i_tu, 2).value,
                 name=tu.cell(6 + i_tu, 3).value,
                 parameter_type='физический',
                 func='C_SC_NA_1 (45)',
                 address=settings.IEC_104['startAddressTU'] + i_tu,
                 description=tu.cell(6 + i_tu, 11).value,
                 channel=str(tu.cell(6 + i_tu, 4).value) + "." + str(tu.cell(6 + i_tu, 5).value))

    """==Физические ТР=="""

    print(' Физические ТР')
    tr = wb.worksheets[4]
    index = index + settings.discrete_output_count

    for i_tr in range(settings.input_count):
        fill_row(bd_sheet=ws_bd,
                 signal_type='ТР',
                 row=index+i_tr,
                 num=tr.cell(6 + i_tr, 2).value,
                 name=tr.cell(6 + i_tr, 3).value,
                 parameter_type='физический',
                 func='C_SE_NC_1 (50)',
                 address=settings.IEC_104['startAddressTRF'] + i_tr,
                 channel=str(tr.cell(6 + i_tr, 4).value) + "." + str(tr.cell(6 + i_tr, 5).value))

    """==Интерфейсные Сигналы=="""

    print(' Интерфейсные ТС')
    index = index + settings.output_count
    for i in range(len(devices.Modbus)):
        ind_tr = 1
        for j in range(len(devices.Modbus[i])):
            transaction = devices.Modbus[i][j]
            if transaction.mfc == 1 or transaction.mfc == 2:
                for k in range(transaction.len_tr):

                    fill_row(bd_sheet=ws_bd,
                             signal_type='ТС',
                             row=index + k,
                             num=ind_tr,
                             name=transaction.name + ' ТC №' + str(k+1),
                             parameter_type='интерфейсный',
                             func='M_SP_NA_1 (1)',
                             address=settings.IEC_104['startAddressTS'] + 200 + k,
                             channel=transaction.name)
                    ind_tr += 1

                index += transaction.len_tr

    print(' Интерфейсные ТИ')
    for i in range(len(devices.Modbus)):
        ind_tr = 1
        for j in range(len(devices.Modbus[i])):
            transaction = devices.Modbus[i][j]
            if transaction.mfc == 3 or transaction.mfc == 4:
                for k in range(transaction.col):

                    fill_row(bd_sheet=ws_bd,
                             signal_type='ТИ',
                             row=index + k,
                             num=ind_tr,
                             name=transaction.name + ' ТИ №' + str(k+1),
                             parameter_type='интерфейсный',
                             func='M_ME_NC_1 (13)',
                             address=settings.IEC_104['startAddressTI'] + 200 + k,
                             channel=transaction.name)
                    ind_tr += 1

                index += transaction.col

    print(' Интерфейсные ТУ')
    for i in range(len(devices.Modbus)):
        ind_tr = 1
        for j in range(len(devices.Modbus[i])):
            transaction = devices.Modbus[i][j]
            if transaction.mfc == 5 or transaction.mfc == 15:
                for k in range(transaction.len_tr):

                    fill_row(bd_sheet=ws_bd,
                             signal_type='ТУ',
                             row=index + k,
                             num=ind_tr,
                             name=transaction.name + ' ТУ №' + str(k+1),
                             parameter_type='интерфейсный',
                             func='C_SC_NA_1 (45)',
                             address=settings.IEC_104['startAddressTU'] + 200 + k,
                             channel=transaction.name)
                    ind_tr += 1

                index += transaction.len_tr

    print(' Интерфейсные ТР')
    for i in range(len(devices.Modbus)):
        ind_tr = 1
        for j in range(len(devices.Modbus[i])):
            transaction = devices.Modbus[i][j]
            if transaction.mfc == 3 or transaction.mfc == 4:
                for k in range(transaction.col):

                    fill_row(bd_sheet=ws_bd,
                             signal_type='ТР',
                             row=index + k,
                             num=ind_tr,
                             name=transaction.name + ' ТР №' + str(k+1),
                             parameter_type='интерфейсный',
                             func='C_SE_NC_1 (50)',
                             address=settings.IEC_104['startAddressTRF'] + 200 + k,
                             channel=transaction.name)
                    ind_tr += 1

                index += transaction.col

    print('База готова')

    """==Сохраняем базу=="""
    wb.save(filename=save_name)
    print('Файл сохранен')

    os.system(save_name)
print("--- %s seconds ---" % (time.time() - start_time))
