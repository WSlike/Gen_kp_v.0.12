import openpyxl
import openfile
from transliterate import translit
from datetime import timedelta

utc3 = timedelta(hours=3)   # часовой пояс +3 часа

"""================Читаем настройки================="""
# Открываем файл excel
print(' Открываем файл...')
filename = openfile.filename
print('================Читаем настройки=================')

filename = filename
wb = openpyxl.load_workbook(filename=filename)

# Когда и кто последний раз редактировал дркумент
date_modified = (wb.properties.modified + utc3).strftime("%Y%m%d_%H%M")
lastModifiedBy = wb.properties.lastModifiedBy
save_name = 'pr_gen_' + date_modified + '_' + translit(lastModifiedBy, reversed=True).replace(' ', '_') + '.xlsx'
print(save_name)

print(' Читаем Количесвто сигналов и Настройки протокола...')
# Считаем количество ТС
ws = wb.worksheets[1]
if ws.cell(2, 2).value is '+':
    discrete_count = ws.max_row - 5
else:
    discrete_count = 0

# Считаем количество ТУ
ws = wb.worksheets[2]
if ws.cell(2, 2).value is '+':
    discrete_output_count = ws.max_row - 5
else:
    discrete_output_count = 0

# Считаем количество ТИ
ws = wb.worksheets[3]
if ws.cell(2, 2).value is '+':
    input_count = ws.max_row - 5
else:
    input_count = 0

# Считаем количество ТР
ws = wb.worksheets[4]
if ws.cell(2, 2).value is '+':
    output_count = ws.max_row - 5
else:
    output_count = 0

# Считаем количество модулей
ws = wb.worksheets[0]
module_count = 0
for i in range(16):
    if ws.cell(10 + i, 4).value is None:
        break
    if module_count < ws.cell(10 + i, 4).value:
        module_count = ws.cell(10 + i, 4).value


# Считываем настройки 104 протокола
ws = wb.worksheets[0]
IEC_104 = {'ASDU': ws.cell(29, 2).value, 'k': ws.cell(30, 2).value, 'w': ws.cell(31, 2).value,
           'timeoutK': ws.cell(32, 2).value, 'startAddressTS': ws.cell(33, 2).value,
           'startAddressTI': ws.cell(34, 2).value, 'startAddressTF': ws.cell(35, 2).value,
           'startAddressTU': ws.cell(36, 2).value, 'startAddressTRI': ws.cell(37, 2).value,
           'startAddressTRF': ws.cell(38, 2).value, 'commandMaxCount': ws.cell(39, 2).value,
           'eventMaxCount': ws.cell(40, 2).value, 'inOutPacketMaxCount': ws.cell(41, 2).value}
wb.close()
