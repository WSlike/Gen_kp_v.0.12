import win32com.client
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename
from openpyxl.reader.excel import load_workbook
from openpyxl import workbook
from transliterate import translit
from datetime import datetime, timedelta
from os import path

try:
    utc3 = timedelta(hours=3)  # часовой пояс +3 часа
    # Открываем окно с выбором файла и запоминаем полное имя
    Tk().withdraw()
    filename = askopenfilename(title="Открыть файл генератора", filetypes=(("Excel files", "*.xlsx"),
                                                                           ("Excel files", "*.xls")))
    # Вытаскиваем путь к файлу
    dir_name = path.dirname(filename) + '/'
    # Инициализируем переменные
    f_open = False
    save_name = ''
    wb = workbook
    # Если выбран какой-то файл, приступаем к обработке
    if filename:
        # Загружаем книгу
        wb = load_workbook(filename=filename)
        # Текущая дата
        date_now = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Проверка на соответствие БД
        worksheets = ['Настройки', 'Карта ТС', 'Карта ТУ', 'Карта ТИ', 'Карта уставок', 'Карта задвижек', 'Карта АПС',
                      'Карта Modbus устройств', 'БД ВУ']
        # Флаг на проверку генератора
        isGen = False
        # Проверяем расположение и названия листов, если все совпадает, то выставляем флаг в 1
        if wb.sheetnames[0] == 'Настройки' and \
                wb.sheetnames[1] == 'Карта ТС' and \
                wb.sheetnames[2] == 'Карта ТУ' and \
                wb.sheetnames[3] == 'Карта ТИ' and \
                wb.sheetnames[4] == 'Карта уставок' and \
                wb.sheetnames[5] == 'Карта задвижек' and \
                wb.sheetnames[6] == 'Карта АПС' and \
                wb.sheetnames[7] == 'Карта Modbus устройств' and \
                wb.sheetnames[8] == 'БД ВУ':
            isGen = True
        else:
            messagebox.showerror("Ошибка", "Выбранный файл не является генератором \n"
                                           "или в нем отсутствуют важные листы")

        # Основная работа

        if isGen:
            # Поднимаем флаг, разрешающий работу с БД
            f_open = True
            # работаем с книгой
            worksheets = wb.sheetnames

            # Когда и кто последний раз редактировал дркумент
            date_modified = (wb.properties.modified + utc3).strftime("%Y%m%d_%H%M%S")
            lastModifiedBy = wb.properties.lastModifiedBy
            wb.worksheets[0].cell(3, 7).value = (wb.properties.modified + utc3).strftime("%d.%m.%Y %H:%M:%S")
            wb.worksheets[0].cell(4, 7).value = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            wb.worksheets[0].cell(5, 7).value = lastModifiedBy

            # Формируем имя документа
            save_name = 'prGen_' + date_modified + '_' + translit(lastModifiedBy, reversed=True).replace(' ',
                                                                                                         '_') + '.xlsx'
            # Проверяем изменялся ли документ
            if save_name not in filename:
                messagebox.showinfo("Генератор",
                                    "Выбранный файл генератора изменен\n"
                                    "Будет создан новый файл\n" +
                                    save_name
                                    )
            # Запускаем СОМ объект Excel
            Excel = win32com.client.Dispatch("Excel.Application")

            # Проверки на открытые документы, так как нельзя сохранить открытый документ после генерирования БД
            if Excel.Workbooks.Count != 0:  # есть открытые Excel документы
                for i in range(1, Excel.Workbooks.Count + 1):
                    if Excel.Workbooks(i).Name in save_name:    # открытый документ совпадает с выбранным

                        if not Excel.Workbooks(i).Saved:
                            """Если выбранный документ открыт, изменен, но не сохранен, 
                            то сохраняем его под новым именем и закрываем"""

                            date_modified = date_now
                            lastModifiedBy = wb.properties.lastModifiedBy
                            save_name = 'prGen_' + date_modified + '_' + \
                                        translit(lastModifiedBy, reversed=True).replace(' ', '_') + '.xlsx'
                            messagebox.showinfo("Генератор",
                                                "Выбранный файл генератора открыт, изменен, но не сохранен\n"
                                                "Будет создан новый файл\n" +
                                                save_name
                                                )
                            filename = Excel.Workbooks(i).Path + save_name
                            Excel.Workbooks(i).SaveAs(filename)
                            Excel.Workbooks(i).Close()
                            wb.close()
                            wb = load_workbook(filename=filename)

                        elif Excel.Workbooks(i).Saved:
                            """Если выбранный документ открыт, и сохранен, то просто закрываем его"""
                            messagebox.showinfo("Генератор",
                                                "Выбранный файл открыт\n"
                                                "БД будет перегенерирована\n" +
                                                save_name
                                                )
                            Excel.Workbooks(i).Close()
                        break

        save_name = dir_name + save_name
    else:
        print('Файл не выбран')
except IOError:
    print("Нельзя работать с открытым файлом!!!")
