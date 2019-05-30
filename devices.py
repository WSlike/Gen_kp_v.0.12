import openpyxl
import openfile


class Transaction:

    def __init__(self, name=None, mfc=None, offset=None, col=None, len_tr=None, timeout=None):
        self.name = name
        self.mfc = mfc
        self.offset = offset
        self.col = col
        self.len_tr = len_tr
        self.timeout = timeout


# Открываем файл excel
print('=======Читаем настройки внешних устройств========')
filename = openfile.filename
wb = openpyxl.load_workbook(filename=openfile.filename)

ws = wb.worksheets[7]
dev = {}  # список {Название устройства: Код устройства}
j = 0
dev_row = []  # массив с начальными строками устройств
for i in range(4, ws.max_row):
    dev_key = ws.cell(i, 4).value
    if dev_key not in dev and (dev_key is not None) and (dev_key != 'Название устройства'):
        dev[dev_key] = ws.cell(i, 3).value
        dev_row.append(i)
dev_row.append(ws.max_row + 2)

print('Используемые устройства:')
for key in dev:
    print(' ' + key)

# Заполнение массива транзакций
Modbus = []
Device = []
for i in range(len(dev_row) - 1):
    first = dev_row[i] + 5
    last = dev_row[i + 1] - 2
    Device = []
    index_tr = 0

    for j in range(first, last+1):
        mfc = ws.cell(j, 6).value  # функция
        offset = ws.cell(j, 7).value  # смещение
        reg_type = ws.cell(j, 8).value  # тип данных
        col = ws.cell(j, 9).value  # количество данных
        timeout = ws.cell(j, 10).value  # таймаут опроса

        if reg_type == 'float' or reg_type == 'long':
            len_tr = col * 2  # количесвто регистров
        else:
            len_tr = col  # количесвто регистров
        t = Transaction(name=ws.cell(dev_row[i], 4).value, mfc=mfc, offset=offset, col=col, len_tr=len_tr, timeout=timeout)
        Device.append(t)
        index_tr += 1

    Modbus.append(Device)

for i in range(len(Modbus)):
    for j in range(len(Modbus[i])):
        transaction = Modbus[i][j]
        # print('\tЗапрос =', j, '\tФункция =', transaction.mfc, '\tсмещение =', transaction.offset,
        #     '\tкол-во регистров =', transaction.len_tr, '\tтаймаут запроса =', transaction.timeout)

wb.close()
